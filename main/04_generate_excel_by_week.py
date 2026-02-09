#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Génère des classeurs Excel hebdomadaires à partir des annonces filtrées.

Le script lit les fichiers JSONL produits par ``03_filter_BODACC_by_day.py``
et regroupe les annonces par semaine ISO à partir de la date figurant dans le
nom du fichier (``YYYYMMDD_bodacc_filtered.jsonl``). Pour chaque semaine,
il crée un fichier Excel dans ``output_dir`` dont le nom commence par
``<année>-W<semaine>`` et contenant une feuille ``BODACC`` avec les colonnes
spécifiées dans la demande métier.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import traceback
from datetime import datetime
from datetime import date
from pathlib import Path
from typing import Dict, List, Sequence
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
import pandas as pd

from utils.utils_get_directories import get_output_dir
from utils.utils_load_config_ini import charger_configuration
from utils.utils_logging import initialiser_logging

COLUMN_MAP: List[tuple[str, Sequence[str]]] = [
    ("ID ANNONCE", ["id"]),
    ("NUMERO ANNONCE", ["numeroannonce"]),
    ("DATE PARUTION", ["dateparution"]),
    (
        "NUMERO_IDENTIFICATION",
        ["listepersonnes/personne/numeroIdentification", "listepersonnes/personne/numeroImmatriculation/numeroIdentification"],
    ),
    ("DENOMINATION", ["listepersonnes/personne/denomination"]),
    ("ACTIVITE", ["listepersonnes/personne/activite"]),
    ("MATRICULE_CCPMA", ["MATRICULE_CCPMA", "MATRICULE_PICRIS_CCPMA"]),
    ("MATRICULE_CPCEA", ["MATRICULE_CPCEA", "MATRICULE_PICRIS_CPCEA"]),
    ("MATRICULE_AGRI", ["MATRICULE_AGRI", "MATRICULE_PICRIS_AGRI"]),
    ("TYPE AVIS", ["typeavis_lib"]),
    ("FAMILLE AVIS", ["familleavis_lib"]),
    ("TYPE JUGEMENT", ["jugement/type"]),
    ("FAMILLE JUGEMENT", ["jugement/famille"]),
    ("NATURE JUGEMENT", ["jugement/nature"]),
    ("DATE JUGEMENT", ["jugement/date"]),
    ("COMPLEMENT JUGEMENT", ["jugement/complementJugement"]),
    ("NOM", ["listepersonnes/personne/nom"]),
    ("PRENOM", ["listepersonnes/personne/prenom"]),
    ("URL", ["url_complete"]),
]

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


from openpyxl import load_workbook
from openpyxl.styles import Font


def fin_semaine_iso(iso_year: int, iso_week: int) -> date:
    """
    Retourne la date du dimanche de la semaine ISO donnée.
    """
    # Lundi de la semaine ISO
    lundi = date.fromisocalendar(iso_year, iso_week, 1)
    # Dimanche
    return lundi + timedelta(days=6)

def convertir_colonne_url_en_hyperliens(fichier_excel: str, feuille: str = "BODACC", colonne_index: int = 19):
    """
    Transforme une colonne contenant des URL en véritables hyperliens Excel.
    colonne_index : 1-based (colonne A=1, B=2, ..., S=19)
    """
    wb = load_workbook(fichier_excel)
    ws = wb[feuille]

    for row in ws.iter_rows(min_row=2):  # on saute l’en-tête
        cell = row[colonne_index - 1]
        url = str(cell.value).strip() if cell.value else ""
        if url.startswith("http"):
            cell.hyperlink = url
            cell.value = "Lien vers annonce"
            cell.style = "Hyperlink"  # Style bleu souligné par défaut
            cell.font = Font(underline="single", color="0563C1")  # en cas de style non pris
        else:
            continue

    wb.save(fichier_excel)
    wb.close()

def mettre_en_forme_excel(
    fichier_excel: str,
    feuille: str = "BODACC",
    auto_largeur: bool = True,
    wrap_text: bool = True
) -> None:
    wb = load_workbook(fichier_excel)
    ws = wb[feuille]

    for row in ws.iter_rows():
        for cell in row:
            # Renvoi à la ligne automatique
            if wrap_text:
                cell.alignment = Alignment(wrapText=True, vertical="top")

    if auto_largeur:
        # Calcul automatique de la largeur des colonnes
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except Exception:
                    pass
            adjusted_width = min((max_length + 2), 80)  # limite max largeur
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(fichier_excel)
    wb.close()


def convertir_feuille_en_table_excel(fichier_excel: Path, feuille: str, nom_table: str = "BODACC"):
    wb = load_workbook(fichier_excel)
    ws = wb[feuille]

    # Détection automatique de la plage (A1 jusqu'à la dernière cellule)
    max_row = ws.max_row
    max_col = ws.max_column
    last_col_letter = ws.cell(row=1, column=max_col).column_letter
    table_range = f"A1:{last_col_letter}{max_row}"

    table = Table(displayName=nom_table, ref=table_range)

    # Style de tableau Excel
    style = TableStyleInfo(
        name="TableStyleMedium9",  # ou TableStyleLight9, Medium2, etc.
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style

    ws.add_table(table)
    wb.save(fichier_excel)
    wb.close()


def _load_bodacc_jsonl(path: Path) -> List[Dict]:
    records: List[Dict] = []
    with path.open("r", encoding="utf-8") as stream:
        for line in stream:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError:
                logging.warning("Ligne JSON invalide ignorée dans %s", path)
    return records


def _parse_day_from_filename(path: Path) -> datetime:
    try:
        day_str = path.stem.split("_")[0]
        return datetime.strptime(day_str, "%Y%m%d")
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Nom de fichier inattendu : {path.name}") from exc


def _collect_weekly_files(filtered_dir: Path) -> Dict[str, List[Path]]:
    weekly_files: Dict[str, List[Path]] = {}
    for jsonl_path in sorted(filtered_dir.glob("*_bodacc_filtered.jsonl")):
        day = _parse_day_from_filename(jsonl_path)
        iso_year, iso_week, _ = day.isocalendar()
        key = f"{iso_year}-W{iso_week:02d}"
        weekly_files.setdefault(key, []).append(jsonl_path)
    return weekly_files


def _collect_values(obj, parts: Sequence[str]) -> List[str]:
    if not parts:
        if obj is None:
            return []
        if isinstance(obj, str):
            return [obj]
        return [str(obj)]

    head, *tail = parts
    if isinstance(obj, list):
        values: List[str] = []
        for item in obj:
            values.extend(_collect_values(item, parts))
        return values

    if isinstance(obj, dict):
        return _collect_values(obj.get(head), tail)

    if isinstance(obj, str):
        # Certaines clés (ex. "jugement", "listepersonnes") sont stockées sous
        # forme de chaîne JSON. On tente un décodage pour accéder aux champs
        # imbriqués.
        try:
            loaded = json.loads(obj)
        except json.JSONDecodeError:
            return []
        return _collect_values(loaded, parts)

    return []


def _extract_field(record: Dict, candidate_paths: Sequence[str]) -> str:
    for path in candidate_paths:
        parts = path.split("/")
        values = _collect_values(record, parts)
        cleaned = [value for value in values if str(value).strip()]
        if cleaned:
            # On dédoublonne en conservant l'ordre d'apparition
            unique_values = list(dict.fromkeys(cleaned))
            return " ; ".join(unique_values)
    return ""


def _build_row(record: Dict) -> Dict[str, str]:
    return {column: _extract_field(record, paths) for column, paths in COLUMN_MAP}


def _ensure_filtered_dir(config) -> Path:
    output_dir = Path(get_output_dir(config))
    filtered_dirname = config["directories"].get("FILTERED_OUTPUT_DIR", "bodacc_filtered_by_day").strip()
    filtered_dir = output_dir / filtered_dirname
    if not filtered_dir.exists() or not filtered_dir.is_dir():
        raise FileNotFoundError(
            f"Répertoire des fichiers filtrés introuvable : {filtered_dir}. "
            "Exécutez d'abord 03_filter_BODACC_by_day pour générer les JSONL."
        )
    return filtered_dir


def _generate_week_excel(week: str, files: List[Path], target_dir: Path) -> None:
    target_path = target_dir / f"{week}_BODACC_DDJC.xlsx"
    if target_path.exists():
        logging.info(f"Fichier déjà présent, génération ignorée : {target_path}")
        return    

    rows: List[Dict[str, str]] = []
    for file in files:
        for record in _load_bodacc_jsonl(file):
            rows.append(_build_row(record))

    df = pd.DataFrame(rows, columns=[col for col, _ in COLUMN_MAP])

    # Écriture initiale
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="BODACC")

    # Conversion en tableau structuré
    convertir_feuille_en_table_excel(fichier_excel=target_path, feuille="BODACC")
    mettre_en_forme_excel(fichier_excel=target_path, feuille="BODACC")
    convertir_colonne_url_en_hyperliens(fichier_excel=target_path, feuille="BODACC", colonne_index=19)

    logging.info("Classeur généré (tableau Excel) : %s (%d lignes)", target_path, len(df))



def main():
    parser = argparse.ArgumentParser(description="Génération d'Excel hebdomadaire BODACC")
    parser.add_argument(
        "--key",
        type=str,
        help="Clé Fernet pour déchiffrer le fichier config.ini (optionnelle)",
    )
    parser.add_argument(
        "--config",
        type=str,
        help="Chemin vers le fichier config.ini (optionnel)",
    )
    args = parser.parse_args()

    config = charger_configuration(config=args.config, key=args.key)
    initialiser_logging(config, log_name="generate_excel_by_day")

    logging.info("===== Début de la génération des Excel hebdomadaires BODACC =====")

    try:
        filtered_dir = _ensure_filtered_dir(config)
        output_dir = Path(get_output_dir(config))

        weekly_files = _collect_weekly_files(filtered_dir)
        if not weekly_files:
            raise FileNotFoundError(
                "Aucun fichier *_bodacc_filtered.jsonl trouvé. "
                "Vérifiez le répertoire ou exécutez l'étape 03."
            )

        aujourdhui = date.today()
        for week, files in sorted(weekly_files.items()):
            iso_year, iso_week = week.split("-W")
            iso_year = int(iso_year)
            iso_week = int(iso_week)

            date_fin_semaine = fin_semaine_iso(iso_year, iso_week)

            if aujourdhui > date_fin_semaine:
                _generate_week_excel(week, files, output_dir)
            else:
                logging.info(
                    "Semaine %s non complète (fin le %s) → génération ignorée",
                    week,
                    date_fin_semaine.isoformat(),
                )


    except Exception:
        logging.error("Erreur critique pendant la génération des Excel BODACC.")
        logging.error(traceback.format_exc())
        sys.exit(1)

    logging.info("===== Fin de la génération des Excel hebdomadaires BODACC =====")


if __name__ == "__main__":
    main()
